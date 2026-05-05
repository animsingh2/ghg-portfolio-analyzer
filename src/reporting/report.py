"""
Module 5 — Reporting Layer
===========================
Assembles outputs from all four modules into two deliverables:

  1.  A structured JSON file — complete, machine-readable record of every
      metric, assumption, and methodology choice.  Suitable for API use,
      audit trails, and feeding other systems.

  2.  A formatted Excel workbook — five sheets mirroring the terminal
      output, formatted for human readers and suitable for sharing with
      investment committees, regulators, or sustainability teams.

      Sheet 1 — Portfolio Summary    (Module 2 headline metrics)
      Sheet 2 — Holdings Detail      (per-holding attribution + emissions)
      Sheet 3 — De-duplication       (Module 3 scope 3 comparison table)
      Sheet 4 — Stress Test          (Module 4a per-asset-class carbon cost)
      Sheet 5 — Pathway Alignment    (Module 4b per-holding alignment)

Usage
-----
    from src.reporting.report import ReportBuilder

    builder = ReportBuilder(
        engine_result=result,
        holding_results=holding_results,
        dedup_result=dedup,
        stress_results=stress_results,
        alignment_result=alignment,
        portfolio=portfolio,
    )

    builder.to_json("reports/PORT_ALPHA_2022.json")
    builder.to_excel("reports/PORT_ALPHA_2022.xlsx")

References
----------
- PCAF Global GHG Accounting & Reporting Standard (2022)
- TCFD Recommendations — Metrics, Targets and Scenario Analysis (2021)
- SFDR Article 29 — Energy & Climate Law (France, 2021)
"""

from __future__ import annotations

import json
import datetime
import logging
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Optional

from src.models import Portfolio, PortfolioEmissionsResult, NGFSScenario
from src.pcaf_engine.engine import HoldingEmissionsResult
from src.deduplication.deduplication import PortfolioDedupResult, DedupMethod
from src.stress_testing.stress_testing import PortfolioStressResult
from src.stress_testing.pathway_alignment import PortfolioAlignmentResult, TemperatureScoreMethod

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Excel style constants
# ---------------------------------------------------------------------------

# Colours (openpyxl uses ARGB hex, no leading #)
C_DARK_GREEN    = "FF1A3A2A"   # Header background
C_MID_GREEN     = "FF2D6B47"   # Section header background
C_LIGHT_GREEN   = "FFE4EFE8"   # Alternate row tint
C_AMBER_LIGHT   = "FFFDF3E0"   # Warning row tint
C_RED_LIGHT     = "FFFBEAE7"   # High-risk row tint
C_WHITE         = "FFFFFFFF"
C_HEADER_TEXT   = "FFFFFFFF"   # White text on dark headers
C_BODY_TEXT     = "FF1A1916"   # Near-black body text
C_MUTED         = "FF6B6960"   # Muted label text
C_BORDER        = "FFD0CEC8"   # Light border


# ---------------------------------------------------------------------------
# Report builder
# ---------------------------------------------------------------------------

class ReportBuilder:
    """
    Assembles all module outputs into JSON and Excel reports.

    All inputs are optional except engine_result and holding_results —
    the report degrades gracefully when downstream modules haven't been run.
    """

    def __init__(
        self,
        engine_result: PortfolioEmissionsResult,
        holding_results: list[HoldingEmissionsResult],
        portfolio: Portfolio,
        dedup_result: Optional[PortfolioDedupResult] = None,
        stress_results: Optional[dict[NGFSScenario, PortfolioStressResult]] = None,
        alignment_result: Optional[PortfolioAlignmentResult] = None,
    ) -> None:
        self.engine = engine_result
        self.holdings = holding_results
        self.portfolio = portfolio
        self.dedup = dedup_result
        self.stress = stress_results
        self.alignment = alignment_result
        self.generated_at = datetime.datetime.utcnow()

    # -------------------------------------------------------------------------
    # JSON export
    # -------------------------------------------------------------------------

    def to_json(self, path: str | Path) -> Path:
        """
        Export the complete report as structured JSON.

        The JSON contains every metric, assumption, methodology choice,
        and per-holding result.  It is designed to be a complete audit
        trail that can reproduce any figure in the Excel report.
        """
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)

        doc = {
            "meta": {
                "generated_at": self.generated_at.isoformat() + "Z",
                "portfolio_id": self.engine.portfolio_id,
                "reporting_year": self.engine.reporting_year,
                "pcaf_standard": "PCAF Global GHG Accounting & Reporting Standard (2022)",
                "gfanz_framework": "GFANZ Transition Finance Metrics (2023)",
            },
            "portfolio_summary": self._json_summary(),
            "holdings": self._json_holdings(),
            "deduplication": self._json_dedup(),
            "stress_test": self._json_stress(),
            "pathway_alignment": self._json_alignment(),
        }

        with open(path, "w", encoding="utf-8") as f:
            json.dump(doc, f, indent=2, default=_json_serialiser)

        logger.info("JSON report written to %s", path)
        return path

    def _json_summary(self) -> dict:
        e = self.engine
        return {
            "total_financed_emissions_tco2e": e.total_financed_emissions_tco2e,
            "scope_1_tco2e": e.scope_1_financed_tco2e,
            "scope_2_tco2e": e.scope_2_financed_tco2e,
            "scope_3_tco2e_raw": e.scope_3_financed_tco2e,
            "scope_3_tco2e_dedup": e.scope_3_financed_tco2e_dedup,
            "waci_tco2e_per_mrevenue": e.waci_tco2e_per_mrevenue,
            "economic_intensity_tco2e_per_musd": e.economic_intensity_tco2e_per_musd,
            "portfolio_weighted_dq_score": e.portfolio_weighted_dq_score,
            "aum_coverage_pct": e.aum_coverage_pct,
            "pct_holdings_with_reported_data": e.pct_holdings_with_reported_data,
            "pct_holdings_estimated": e.pct_holdings_estimated,
            "n_holdings_total": e.n_holdings_total,
            "n_holdings_with_emissions": e.n_holdings_with_emissions,
            "total_aum_usd": self.portfolio.total_aum_usd,
            "by_asset_class": e.by_asset_class,
        }

    def _json_holdings(self) -> list[dict]:
        return [
            {
                "holding_id": h.holding_id,
                "entity_name": h.entity_name,
                "asset_class": h.asset_class.value,
                "outstanding_amount_usd": h.outstanding_amount_usd,
                "attribution_factor": h.attribution_factor,
                "financed_scope_1_tco2e": h.financed_scope_1_tco2e,
                "financed_scope_2_tco2e": h.financed_scope_2_tco2e,
                "financed_scope_3_tco2e": h.financed_scope_3_tco2e,
                "financed_total_tco2e": h.financed_total_tco2e,
                "weighted_dq_score": h.weighted_dq_score,
                "error_margin_pct": h.error_margin_pct,
                "waci_contribution": h.waciContrib if hasattr(h, 'waciContrib') else None,
            }
            for h in self.holdings
        ]

    def _json_dedup(self) -> Optional[dict]:
        if not self.dedup:
            return None
        d = self.dedup
        return {
            "method": d.method.value,
            "multiplier_used": d.multiplier_used,
            "scope_3_raw_tco2e": d.scope_3_raw_tco2e,
            "scope_3_dedup_tco2e": d.scope_3_dedup_tco2e,
            "estimated_double_count_tco2e": d.estimated_double_count_tco2e,
            "double_count_pct_of_raw": d.double_count_pct_of_raw,
            "total_raw_tco2e": d.total_raw_tco2e,
            "total_dedup_tco2e": d.total_dedup_tco2e,
            "diagnostics": d.multiplier_diagnostics,
        }

    def _json_stress(self) -> Optional[dict]:
        if not self.stress:
            return None
        out = {}
        for scenario, r in self.stress.items():
            out[scenario.value] = {
                "scenario_label": r.scenario_label,
                "horizon_year": r.horizon_year,
                "carbon_price_usd_per_tco2e": r.carbon_price_usd_per_tco2e,
                "total_carbon_cost_usd": r.total_carbon_cost_usd,
                "portfolio_carbon_cost_pct_revenue": r.portfolio_carbon_cost_pct_revenue,
                "n_high_risk": r.n_high_risk,
                "n_medium_risk": r.n_medium_risk,
                "n_low_risk": r.n_low_risk,
                "high_risk_aum_pct": r.high_risk_aum_pct,
                "by_asset_class": r.by_asset_class,
                "high_risk_holdings": [
                    {
                        "holding_id": h.holding_id,
                        "entity_name": h.entity_name,
                        "carbon_cost_pct_revenue": h.carbon_cost_pct_revenue,
                        "implied_carbon_cost_usd": h.implied_carbon_cost_usd,
                    }
                    for h in r.holding_results if h.is_high_risk
                ],
            }
        return out

    def _json_alignment(self) -> Optional[dict]:
        if not self.alignment:
            return None
        a = self.alignment
        return {
            "assessment_year": a.assessment_year,
            "target_year": a.target_year,
            "temperature_method": a.temperature_method,
            "implied_temperature_no_cap": a.implied_temperature_no_cap,
            "implied_temperature_capped": a.implied_temperature_capped,
            "implied_temperature_winsorised": a.implied_temperature_winsorised,
            "implied_temperature_primary": a.implied_temperature_c,
            "pct_aum_aligned": a.pct_aum_aligned,
            "pct_aum_ahead_of_path": a.pct_aum_ahead_of_path,
            "pct_aum_misaligned": a.pct_aum_misaligned,
            "pct_aum_no_data": a.pct_aum_no_data,
            "total_financed_gap_tco2e": a.total_financed_gap_tco2e,
            "top_misaligned": a.top_misaligned,
            "holdings": [
                {
                    "holding_id": h.holding_id,
                    "entity_name": h.entity_name,
                    "sector": h.sector,
                    "alignment_status": h.alignment_status,
                    "entity_intensity_tco2e_per_mrevenue": h.entity_intensity,
                    "benchmark_intensity_tco2e_per_mrevenue": h.benchmark_intensity,
                    "intensity_gap_pct": h.intensity_gap_pct,
                    "required_annual_reduction_pct": h.required_annual_reduction_pct,
                    "financed_gap_tco2e": h.financed_gap_tco2e,
                }
                for h in a.holding_results
            ],
        }

    # -------------------------------------------------------------------------
    # Excel export
    # -------------------------------------------------------------------------

    def to_excel(self, path: str | Path) -> Path:
        """
        Export the complete report as a formatted Excel workbook.

        Five sheets, one per module output.  Formatted for human readers
        with colour-coded headers, conditional row highlighting, and
        consistent number formats throughout.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side, numbers
            )
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Run: pip install openpyxl")

        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        wb.remove(wb.active)  # remove default empty sheet

        self._sheet_summary(wb)
        self._sheet_holdings(wb)
        if self.dedup:
            self._sheet_dedup(wb)
        if self.stress:
            self._sheet_stress(wb)
        if self.alignment:
            self._sheet_alignment(wb)

        wb.save(path)
        logger.info("Excel report written to %s", path)
        return path

    # ── Sheet helpers ──────────────────────────────────────────────────────

    def _sheet_summary(self, wb) -> None:
        from openpyxl.styles import Font, PatternFill, Alignment
        ws = wb.create_sheet("1. Portfolio Summary")
        ws.sheet_view.showGridLines = False

        e = self.engine
        port = self.portfolio

        _title(ws, "GHG FINANCED EMISSIONS — PORTFOLIO SUMMARY", 1, 1, span=4)
        _subtitle(ws, f"Portfolio: {e.portfolio_id}  |  Reporting year: {e.reporting_year}  |  "
                  f"Generated: {self.generated_at.strftime('%Y-%m-%d')}", 2, 1, span=4)

        r = 4
        _section(ws, "ABSOLUTE FINANCED EMISSIONS", r, 1, span=4); r += 1
        rows = [
            ("Scope 1 (direct emissions)", e.scope_1_financed_tco2e, "tCO₂e"),
            ("Scope 2 (purchased energy)", e.scope_2_financed_tco2e, "tCO₂e"),
            ("Scope 3 (value chain) — raw", e.scope_3_financed_tco2e, "tCO₂e"),
            ("Scope 3 (value chain) — de-duplicated",
             getattr(e, 'scope_3_financed_tco2e_dedup', None), "tCO₂e"),
            ("TOTAL financed emissions", e.total_financed_emissions_tco2e, "tCO₂e"),
        ]
        for i, (label, val, unit) in enumerate(rows):
            bold = label.startswith("TOTAL")
            bg = C_LIGHT_GREEN if i % 2 == 0 else C_WHITE
            _data_row(ws, r, [label, _n(val), unit], bold=bold, bg=bg)
            r += 1

        r += 1
        _section(ws, "INTENSITY METRICS", r, 1, span=4); r += 1
        _data_row(ws, r, ["WACI", _n(e.waci_tco2e_per_mrevenue, 2),
                          "tCO₂e / $M revenue",
                          "Weighted Average Carbon Intensity"], bg=C_LIGHT_GREEN); r += 1
        _data_row(ws, r, ["Economic intensity",
                          _n(e.economic_intensity_tco2e_per_musd, 2),
                          "tCO₂e / $M AUM",
                          "Total financed emissions / AUM"]); r += 1
        _data_row(ws, r, ["Total AUM",
                          _n(port.total_aum_usd / 1e6, 1),
                          "$M USD"], bg=C_LIGHT_GREEN); r += 1

        r += 1
        _section(ws, "DATA QUALITY", r, 1, span=4); r += 1
        _data_row(ws, r, ["Portfolio weighted DQ score",
                          _n(e.portfolio_weighted_dq_score, 2),
                          "/ 5  (1=verified, 5=estimated)"], bg=C_LIGHT_GREEN); r += 1
        _data_row(ws, r, ["AUM coverage",
                          _pct(e.aum_coverage_pct), "%",
                          "% of AUM with matched emissions data"]); r += 1
        _data_row(ws, r, ["Holdings with reported data",
                          _pct(e.pct_holdings_with_reported_data), "%",
                          "DQ score 1 or 2"], bg=C_LIGHT_GREEN); r += 1
        _data_row(ws, r, ["Holdings estimated",
                          _pct(e.pct_holdings_estimated), "%",
                          "DQ score 4 or 5"]); r += 1

        r += 1
        _section(ws, "BY ASSET CLASS", r, 1, span=4); r += 1
        _header_row(ws, r, ["Asset Class", "AUM ($M)", "Financed Emissions (tCO₂e)", ""]); r += 1
        for i, (ac, data) in enumerate(
            sorted(e.by_asset_class.items(),
                   key=lambda x: x[1].get("financed_emissions_tco2e", 0), reverse=True)
        ):
            bg = C_LIGHT_GREEN if i % 2 == 0 else C_WHITE
            _data_row(ws, r, [
                ac.replace("_", " ").title(),
                _n(data.get("aum_usd", 0) / 1e6, 1),
                _n(data.get("financed_emissions_tco2e", 0)),
                ""
            ], bg=bg); r += 1

        _autofit(ws, [45, 14, 28, 40])

    def _sheet_holdings(self, wb) -> None:
        ws = wb.create_sheet("2. Holdings Detail")
        ws.sheet_view.showGridLines = False

        _title(ws, "HOLDINGS — FINANCED EMISSIONS DETAIL", 1, 1, span=8)
        _subtitle(ws, f"Portfolio: {self.engine.portfolio_id}  |  "
                  f"Reporting year: {self.engine.reporting_year}", 2, 1, span=8)

        r = 4
        headers = ["Entity", "Asset Class", "Outstanding ($M)",
                   "Attribution %", "Scope 1+2 (tCO₂e)", "Scope 3 (tCO₂e)",
                   "Total (tCO₂e)", "DQ Score"]
        _header_row(ws, r, headers); r += 1

        sorted_holdings = sorted(
            self.holdings,
            key=lambda h: h.financed_total_tco2e or 0, reverse=True
        )
        for i, h in enumerate(sorted_holdings):
            dq = h.weighted_dq_score or 0
            if dq >= 4:
                bg = C_AMBER_LIGHT
            elif i % 2 == 0:
                bg = C_LIGHT_GREEN
            else:
                bg = C_WHITE

            s12 = (h.financed_scope_1_tco2e or 0) + (h.financed_scope_2_tco2e or 0)
            _data_row(ws, r, [
                h.entity_name,
                h.asset_class.value.replace("_", " ").title(),
                _n(h.outstanding_amount_usd / 1e6, 2),
                _pct(h.attribution_factor) if h.attribution_factor else "—",
                _n(s12),
                _n(h.financed_scope_3_tco2e),
                _n(h.financed_total_tco2e),
                _n(h.weighted_dq_score, 1) if h.weighted_dq_score else "—",
            ], bg=bg); r += 1

        _autofit(ws, [35, 30, 16, 14, 20, 18, 18, 10])

    def _sheet_dedup(self, wb) -> None:
        ws = wb.create_sheet("3. Scope 3 De-duplication")
        ws.sheet_view.showGridLines = False
        d = self.dedup

        _title(ws, "SCOPE 3 DE-DUPLICATION", 1, 1, span=4)
        _subtitle(ws, f"Method: {d.method.value.replace('_',' ').title()}  |  "
                  f"Multiplier: {d.multiplier_used:.4f}", 2, 1, span=4)

        r = 4
        _section(ws, "METHODOLOGY", r, 1, span=4); r += 1
        diag = d.multiplier_diagnostics
        _data_row(ws, r, ["Method", d.method.value.replace("_", " ").title()], bg=C_LIGHT_GREEN); r += 1
        _data_row(ws, r, ["Multiplier applied", f"{d.multiplier_used:.4f}"]); r += 1
        if "overlap_score" in diag:
            _data_row(ws, r, ["Overlap score", f"{diag['overlap_score']:.4f}",
                              "", "Estimated fraction of Scope 3 double-counted"],
                      bg=C_LIGHT_GREEN); r += 1
        if "vs_msci" in diag:
            _data_row(ws, r, ["vs MSCI 0.205", f"{diag['vs_msci']:+.4f}",
                              "", "Positive = less de-duplication than MSCI would apply"]); r += 1
        if "interpretation" in diag:
            _data_row(ws, r, ["Interpretation", diag["interpretation"]],
                      bg=C_LIGHT_GREEN); r += 1

        r += 1
        _section(ws, "RESULTS", r, 1, span=4); r += 1
        _header_row(ws, r, ["Approach", "Scope 3 (tCO₂e)", "Total S1+2+3 (tCO₂e)", ""]); r += 1
        _data_row(ws, r, ["Raw (no de-duplication)",
                          _n(d.scope_3_raw_tco2e), _n(d.total_raw_tco2e), ""],
                  bg=C_LIGHT_GREEN); r += 1
        _data_row(ws, r, [f"Adjusted (×{d.multiplier_used:.4f})",
                          _n(d.scope_3_dedup_tco2e), _n(d.total_dedup_tco2e), ""]); r += 1
        _data_row(ws, r, ["Double-count removed",
                          _n(d.estimated_double_count_tco2e),
                          f"{d.double_count_pct_of_raw*100:.1f}% of raw Scope 3", ""],
                  bg=C_LIGHT_GREEN); r += 1

        if "sector_weights" in diag and diag["sector_weights"]:
            r += 1
            _section(ws, "SECTOR WEIGHTS (by Scope 3 emissions)", r, 1, span=4); r += 1
            _header_row(ws, r, ["Sector", "Weight (%)", "", ""]); r += 1
            for i, (sector, weight) in enumerate(diag["sector_weights"].items()):
                bg = C_LIGHT_GREEN if i % 2 == 0 else C_WHITE
                _data_row(ws, r, [sector, f"{weight*100:.1f}%", "", ""], bg=bg); r += 1

        _autofit(ws, [40, 22, 24, 50])

    def _sheet_stress(self, wb) -> None:
        ws = wb.create_sheet("4. Stress Test")
        ws.sheet_view.showGridLines = False

        first = next(iter(self.stress.values()))
        _title(ws, "TRANSITION RISK — CARBON PRICE STRESS TEST", 1, 1, span=6)
        _subtitle(ws, f"Portfolio: {first.portfolio_id}  |  Horizon year: {first.horizon_year}",
                  2, 1, span=6)

        r = 4
        _section(ws, "SCENARIO COMPARISON", r, 1, span=6); r += 1
        _header_row(ws, r, ["Scenario", "Carbon Price ($/tCO₂e)",
                             "Total Carbon Cost", "High Risk Holdings",
                             "Medium Risk", "Low Risk"]); r += 1
        for i, (scenario, res) in enumerate(self.stress.items()):
            bg = C_LIGHT_GREEN if i % 2 == 0 else C_WHITE
            _data_row(ws, r, [
                res.scenario_label,
                f"${res.carbon_price_usd_per_tco2e:.0f}",
                f"${res.total_carbon_cost_usd/1e6:.1f}M",
                str(res.n_high_risk),
                str(res.n_medium_risk),
                str(res.n_low_risk),
            ], bg=bg); r += 1

        # Asset class breakdown — Orderly scenario
        orderly = self.stress.get(NGFSScenario.ORDERLY, first)
        r += 1
        _section(ws, f"BY ASSET CLASS  [{orderly.scenario_label}]", r, 1, span=6); r += 1
        _header_row(ws, r, ["Asset Class", "AUM ($M)", "Carbon Cost",
                             "% of Revenue", "High Risk Holdings", ""]); r += 1
        for i, (ac, bucket) in enumerate(
            sorted(orderly.by_asset_class.items(),
                   key=lambda x: x[1].get("carbon_cost_usd", 0), reverse=True)
        ):
            bg = C_LIGHT_GREEN if i % 2 == 0 else C_WHITE
            pct_rev = bucket.get("pct_revenue")
            is_sovereign = ac == "sovereign_debt"
            _data_row(ws, r, [
                ac.replace("_", " ").title(),
                _n(bucket.get("aum_usd", 0) / 1e6, 1),
                f"${bucket.get('carbon_cost_usd', 0)/1e6:.2f}M",
                "See note *" if is_sovereign else (_pct_str(pct_rev) if pct_rev else "N/A"),
                str(bucket.get("n_high_risk", 0)) if bucket.get("n_high_risk", 0) > 0 else "—",
                "",
            ], bg=bg); r += 1

        if orderly.has_sovereign_debt:
            _data_row(ws, r, ["* Sovereign debt % revenue uses government tax receipts — "
                               "not comparable to corporate holdings.", "", "", "", "", ""],
                      bg=C_AMBER_LIGHT); r += 1

        # High-risk holdings across all scenarios
        all_high = {
            h.holding_id
            for res in self.stress.values()
            for h in res.holding_results if h.is_high_risk
        }
        if all_high:
            r += 1
            _section(ws, "HIGH-RISK HOLDINGS (carbon cost > 5% of revenue)", r, 1, span=6); r += 1
            _header_row(ws, r, ["Company", "Asset Class",
                                 "Orderly", "Disorderly", "Hot House", ""]); r += 1
            for hid in sorted(all_high):
                row_vals = []
                name = ac_label = ""
                for scenario in NGFSScenario:
                    res = self.stress.get(scenario)
                    if res:
                        h = next((x for x in res.holding_results if x.holding_id == hid), None)
                        if h:
                            name = h.entity_name
                            ac_label = h.asset_class.value.replace("_", " ").title()
                            pct = h.carbon_cost_pct_revenue
                            row_vals.append(f"{'▲ ' if h.is_high_risk else ''}"
                                            f"{pct*100:.1f}%" if pct else "—")
                        else:
                            row_vals.append("—")
                bg = C_RED_LIGHT
                _data_row(ws, r, [name, ac_label] + row_vals + [""], bg=bg); r += 1

        _autofit(ws, [38, 20, 16, 16, 20, 10])

    def _sheet_alignment(self, wb) -> None:
        ws = wb.create_sheet("5. Pathway Alignment")
        ws.sheet_view.showGridLines = False
        a = self.alignment

        _title(ws, "SBTi 1.5°C PATHWAY ALIGNMENT", 1, 1, span=6)
        _subtitle(ws, f"Assessment year: {a.assessment_year}  |  Target year: {a.target_year}",
                  2, 1, span=6)

        r = 4
        _section(ws, "IMPLIED TEMPERATURE SCORE", r, 1, span=6); r += 1
        temp_rows = [
            ("No cap (raw)", a.implied_temperature_no_cap,
             "Academically pure — can be distorted by outliers"),
            (f"Overshoot cap ({a.overshoot_cap_used:.0f}× benchmark)  [MSCI standard]",
             a.implied_temperature_capped,
             "Most widely used — caps each holding at 10× benchmark overshoot"),
            (f"Winsorised ({a.winsorise_pct_used*100:.0f}th percentile)",
             a.implied_temperature_winsorised,
             "Data-driven cap — uses portfolio's own distribution"),
        ]
        method_map = {
            TemperatureScoreMethod.NO_CAP: 0,
            TemperatureScoreMethod.OVERSHOOT_CAP: 1,
            TemperatureScoreMethod.WINSORISE: 2,
        }
        primary_idx = method_map.get(a.temperature_method, 1)
        for i, (label, temp, note) in enumerate(temp_rows):
            is_primary = (i == primary_idx)
            bg = C_LIGHT_GREEN if is_primary else C_WHITE
            temp_str = f"~{temp:.1f}°C" if temp is not None else "N/A"
            primary_str = "◀ PRIMARY" if is_primary else ""
            _data_row(ws, r, [label, temp_str, primary_str, note], bg=bg); r += 1

        r += 1
        _section(ws, "PORTFOLIO ALIGNMENT SUMMARY", r, 1, span=6); r += 1
        summary_rows = [
            ("AUM aligned with 1.5°C", _pct(a.pct_aum_aligned), "%"),
            ("AUM ahead of pathway", _pct(a.pct_aum_ahead_of_path), "%"),
            ("AUM misaligned", _pct(a.pct_aum_misaligned), "%"),
            ("AUM — no data", _pct(a.pct_aum_no_data), "%"),
            ("Total financed gap", _n(a.total_financed_gap_tco2e), "tCO₂e"),
        ]
        for i, (label, val, unit) in enumerate(summary_rows):
            bg = C_LIGHT_GREEN if i % 2 == 0 else C_WHITE
            _data_row(ws, r, [label, val, unit, ""], bg=bg); r += 1

        r += 1
        _section(ws, "HOLDING DETAIL", r, 1, span=6); r += 1
        _header_row(ws, r, ["Company", "Sector", "Status",
                             "Gap vs 1.5°C", "Req. Annual Reduction",
                             "Financed Gap (tCO₂e)"]); r += 1

        STATUS_ORDER = {"misaligned": 0, "aligned": 1, "ahead_of_path": 2, "no_data": 3}
        for h in sorted(a.holding_results,
                        key=lambda x: (STATUS_ORDER.get(x.alignment_status, 9),
                                       -(x.intensity_gap or 0))):
            if h.alignment_status == "misaligned":
                bg = C_RED_LIGHT
            elif h.alignment_status == "ahead_of_path":
                bg = C_LIGHT_GREEN
            else:
                bg = C_WHITE

            status_labels = {
                "aligned": "✓ Aligned",
                "ahead_of_path": "★ Ahead of path",
                "misaligned": "✗ Misaligned",
                "no_data": "○ No data",
            }
            gap = (f"{h.intensity_gap_pct*100:+.0f}%"
                   if h.intensity_gap_pct is not None else "—")
            req = (f"{h.required_annual_reduction_pct*100:.1f}%/yr"
                   if h.required_annual_reduction_pct else "—")
            _data_row(ws, r, [
                h.entity_name,
                h.sector or "—",
                status_labels.get(h.alignment_status, h.alignment_status),
                gap,
                req,
                _n(h.financed_gap_tco2e) if h.financed_gap_tco2e else "—",
            ], bg=bg); r += 1

        _autofit(ws, [35, 25, 16, 14, 22, 22])


# ---------------------------------------------------------------------------
# Openpyxl helpers
# ---------------------------------------------------------------------------

def _title(ws, text, row, col, span=4):
    from openpyxl.styles import Font, PatternFill, Alignment
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Arial", size=13, bold=True, color=C_HEADER_TEXT)
    cell.fill = PatternFill("solid", fgColor=C_DARK_GREEN)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)


def _subtitle(ws, text, row, col, span=4):
    from openpyxl.styles import Font, PatternFill, Alignment
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Arial", size=10, color=C_HEADER_TEXT)
    cell.fill = PatternFill("solid", fgColor=C_MID_GREEN)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 16
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)


def _section(ws, text, row, col, span=4):
    from openpyxl.styles import Font, PatternFill, Alignment
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Arial", size=10, bold=True, color=C_HEADER_TEXT)
    cell.fill = PatternFill("solid", fgColor=C_MID_GREEN)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 16
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)


def _header_row(ws, row, values):
    from openpyxl.styles import Font, PatternFill, Alignment
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(name="Arial", size=9, bold=True, color=C_HEADER_TEXT)
        cell.fill = PatternFill("solid", fgColor="FF3D6B52")
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 15


def _data_row(ws, row, values, bold=False, bg=C_WHITE):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    border_side = Side(style="thin", color=C_BORDER)
    bottom_border = Border(bottom=border_side)
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(name="Arial", size=9, bold=bold, color=C_BODY_TEXT)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                   indent=1, wrap_text=True)
        cell.border = bottom_border
    ws.row_dimensions[row].height = 15


def _autofit(ws, widths: list[float]):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _n(val, decimals=0) -> str:
    if val is None:
        return "—"
    if decimals == 0:
        return f"{val:,.0f}"
    return f"{val:,.{decimals}f}"


def _pct(val) -> str:
    if val is None:
        return "—"
    return f"{val*100:.1f}"


def _pct_str(val) -> str:
    if val is None:
        return "—"
    return f"{val*100:.1f}%"


def _json_serialiser(obj):
    """Handle non-serialisable types for json.dump."""
    if hasattr(obj, 'isoformat'):
        return obj.isoformat()
    if hasattr(obj, 'value'):
        return obj.value
    return str(obj)
